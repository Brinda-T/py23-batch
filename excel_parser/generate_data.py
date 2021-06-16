from openpyxl import load_workbook

from epics_stories_data import( get_issue_key_pos,
get_issue_id_pos,
get_custom_field_link_pos,
get_ename_pos,
get_assignee_pos,
get_orig_esti_pos, 
get_progress_pos,
get_custom_field_points_pos,
get_spenthours_pos,
get_rem_esti_pos,
get_esdate_pos,
get_etdate_pos,
get_easdate_pos, 
get_eaedate_pos, 
get_originalhours_pos,
get_remaininghours_pos,
get_time_spent_pos,
get_sprint_pos,
get_sprint2_pos,
get_sprint3_pos,
get_summary_pos,
get_sno_pos,
get_epic_custom_field_link,
get_epic_ename,
get_epic_esdate,
get_epic_etdate,
get_epic_easDate,
get_epic_eaedate,
get_epic_id_col,
get_epic_id_npos,
get_story_id_col,
get_story_id_npos,
get_sprint_id_col,
get_sprint_id_npos,
get_story_desc_col,
get_story_desc_npos,
get_assigned_to_col,
get_assigned_to_npos,
get_esti_story_points_col,
get_esti_story_points_npos,
get_story_planned_start_col,
get_story_planned_start_npos,
get_story_planned_end_col,
get_story_planned_end_npos,
get_esti_effort_in_hrs_col,
get_esti_effort_in_hrs_npos,
get_time_spent_in_sec_col,
get_time_spent_in_sec_npos,
get_effort_cons_in_hrs_col,
get_effort_cons_in_hrs_npos,
get_pending_effort_in_hrs_col,
get_pending_effort_in_hrs_npos,
get_story_actual_start_col,
get_story_actual_start_npos,
get_story_actual_end_col,
get_story_actual_end_npos,
get_story_effort_comp_col,
get_story_effort_comp_npos,
get_story_schedule_progress_pos,
get_story_schedule_progress_col,
get_story_schedule_progress_npos,
get_story_schedule_overrun_pos,
get_story_schedule_overrun_col,
get_story_schedule_overrun_npos,
get_remarks_pos,
get_remarks_col,
get_remarks_npos,

)

ik_pos = get_issue_key_pos()
ii_pos = get_issue_id_pos()
cfl_pos = get_custom_field_link_pos()
en_pos = get_ename_pos()
a_pos = get_assignee_pos()
oe_pos = get_orig_esti_pos() 
p_pos = get_progress_pos()
cfp_pos = get_custom_field_points_pos()
sh_pos = get_spenthours_pos()
re_pos = get_rem_esti_pos()
es_pos = get_esdate_pos()
et_pos = get_etdate_pos()
eas_pos = get_easdate_pos() 
eae_pos = get_eaedate_pos()
oh_pos = get_originalhours_pos()
rh_pos = get_remaininghours_pos()
ts_pos = get_time_spent_pos()
s1_pos = get_sprint_pos()
s2_pos = get_sprint2_pos()
s3_pos = get_sprint3_pos()
s_pos = get_summary_pos()
sno_pos = get_sno_pos()
epic_cfl = get_epic_custom_field_link()
epic_en_pos = get_epic_ename()
epic_es_pos = get_epic_esdate()
epic_et_pos = get_epic_etdate()
epic_eas_pos = get_epic_easDate()
epic_eae_pos = get_epic_eaedate()
ei_col = get_epic_id_col()
ei_npos = get_epic_id_npos()
si_col = get_story_id_col()
si_npos = get_story_id_npos()
sprint_id_col = get_sprint_id_col()
sprint_id_npos = get_sprint_id_npos()
sd_col = get_story_desc_col()
sd_npos = get_story_desc_npos()
at_col = get_assigned_to_col()
at_npos = get_assigned_to_npos()
esp_col = get_esti_story_points_col()
esp_npos = get_esti_story_points_npos()
sps_col = get_story_planned_start_col()
sps_npos = get_story_planned_start_npos()
spe_col = get_story_planned_end_col()
spe_npos = get_story_planned_end_npos()
ee_in_hrs_col = get_esti_effort_in_hrs_col()
ee_in_hrs_npos = get_esti_effort_in_hrs_npos()
ec_in_hrs_col = get_effort_cons_in_hrs_col()
ec_in_hrs_npos = get_effort_cons_in_hrs_npos()
ts_in_sec_col = get_time_spent_in_sec_col()
ts_in_sec_npos = get_time_spent_in_sec_npos()
pe_in_hrs_col = get_pending_effort_in_hrs_col()
pe_in_hrs_npos = get_pending_effort_in_hrs_npos()
sas_col = get_story_actual_start_col()
sas_npos = get_story_actual_start_npos()
sae_col = get_story_actual_end_col()
sae_npos = get_story_actual_end_npos()
sec_col = get_story_effort_comp_col()
sec_npos = get_story_effort_comp_npos()
ssp_pos = get_story_schedule_progress_pos()
ssp_col = get_story_schedule_progress_col()
ssp_npos = get_story_schedule_progress_npos()
sso_pos = get_story_schedule_overrun_pos()
sso_col = get_story_schedule_overrun_col()
sso_npos = get_story_schedule_overrun_npos()
r_pos = get_remarks_pos()
r_col = get_remarks_col()
r_npos = get_remarks_npos()
def append_rows_by_name(wbook, dst_wname, src1_wname, src2_wname): 
    wb = load_workbook(wbook)
    swsheet1 = wb.get_sheet_by_name(src1_wname)
    swsheet2 = wb.get_sheet_by_name(src2_wname)

    try:
        dwsheet = wb.get_sheet_by_name(dst_wname)
        print("Worksheet '%s' found" %(dst_wname))

        print("Removing worksheet: '%s'" %(dst_wname))
        wb.remove_sheet(dst_wname)
    except:
        print("Worksheet '%s' not found" %(dst_wname))
    finally:
        print("Creating new worksheet : '%s'" %(dst_wname))
        dwsheet = wb.create_sheet(dst_wname)

    srcount1 = swsheet1.max_row
    sccount1 = swsheet1.max_column

    srcount2 = swsheet2.max_row
    sccount2 = swsheet2.max_column

    drcount = dwsheet.max_row
    dccount = dwsheet.max_column


    for row in swsheet2.iter_rows():
        if(row[2].value != None):
            frow = [cell.value for cell in row]
            dwsheet.append(frow)
    
    dwsheet.insert_cols(es_pos)
    dwsheet.insert_cols(et_pos)
    dwsheet.insert_cols(eas_pos)
    dwsheet.insert_cols(eae_pos)
    dwsheet.insert_cols(p_pos+1)
    dwsheet.insert_cols(oh_pos+1)
    dwsheet.insert_cols(sh_pos+1)
    dwsheet.insert_cols(rh_pos+1)
	
    dwsheet[1][p_pos].value = "progress"
    dwsheet[1][oh_pos].value = "originalhours"
    dwsheet[1][sh_pos].value = "spenthours"
    dwsheet[1][rh_pos].value = "remaininghours"
    
    for i in range(2,srcount2+1):
        dwsheet[i][oh_pos].value = round(int(dwsheet[i][12].value)/3600)
        dwsheet[i][sh_pos].value = round(int(dwsheet[i][14].value)/3600)
        dwsheet[i][rh_pos].value = round(int(dwsheet[i][16].value)/3600)
        dwsheet[i][p_pos].value = round(int(dwsheet[i][15].value)/int(dwsheet[i][13].value)*100)

    for i in range(1, srcount1+1):
        print(swsheet1[i][1].value)
        for j in range(1,srcount2+1):
            if(swsheet1[i][1].value == swsheet2[j][3].value):
                
                dwsheet[j][4].value = swsheet1[i][3].value
                dwsheet[j][5].value = swsheet1[i][4].value
                dwsheet[j][6].value = swsheet1[i][5].value
                dwsheet[j][7].value = swsheet1[i][6].value
 

    dwsheet[1][cfl_pos].value = ei_col
    dwsheet[1][ik_pos].value = si_col
    dwsheet[1][s1_pos].value = sprint_id_col
    dwsheet[1][s_pos].value = sd_col
    dwsheet[1][a_pos].value = at_col
    dwsheet[1][cfp_pos].value = esp_col
    dwsheet[1][es_pos-1].value = sps_col
    dwsheet[1][et_pos-1].value = spe_col
    dwsheet[1][sh_pos].value = ee_in_hrs_col
    dwsheet[1][oe_pos].value = ts_in_sec_col
    dwsheet[1][oh_pos].value = ec_in_hrs_col
    dwsheet[1][rh_pos].value = pe_in_hrs_col
    dwsheet[1][eas_pos-1].value = sas_col
    dwsheet[1][eae_pos-1].value = sae_col
    dwsheet[1][p_pos].value = sec_col
    dwsheet.insert_cols(ssp_pos)
    dwsheet.insert_cols(sso_pos)
    dwsheet.insert_cols(r_pos)
    dwsheet[1][ssp_pos-1].value = ssp_col
    dwsheet[1][sso_pos-1].value = sso_col
    dwsheet[1][r_pos-1].value = r_col
    for j in range (2, int(srcount2) + 1):
        
        dwsheet[j][ssp_pos-1].value = "100%"
        dwsheet[j][sso_pos-1].value = "0%"
         
                


    wb.save(wbook)


def main():
    epic_data = "01-epics"
    stories_data = "02-stories"
    wbook = "generate-files.xlsx"
    dst_wname = "epic_stories"

    append_rows_by_name(wbook, dst_wname, epic_data, stories_data)

if (__name__ == "__main__"):
    main()
