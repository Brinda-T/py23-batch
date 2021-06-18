from openpyxl import load_workbook

from epics_stories_data import( get_issue_key_pos,
get_issue_id_pos,
get_custom_field_link_pos,
get_ename_pos,
get_assignee_pos,
get_orig_esti_pos, 
get_custom_field_points_pos,
get_rem_esti_pos,
get_time_spent_pos,
get_sprint_pos,
get_sprint2_pos,
get_sprint3_pos,
get_summary_pos,
get_sno_pos,
get_epic_custom_field_link,
get_epic_ename,
get_epic_esdate,
get_epic_etdate,
get_epic_easDate,
get_epic_eaedate,
get_epic_id_col,
get_epic_id_npos,
get_story_id_col,
get_story_id_npos,
get_sprint_id_col,
get_sprint_id_npos,
get_story_desc_col,
get_story_desc_npos,
get_assigned_to_col,
get_assigned_to_npos,
get_esti_story_points_col,
get_esti_story_points_npos,
get_story_planned_start_col,
get_story_planned_start_npos,
get_story_planned_end_col,
get_story_planned_end_npos,
get_esti_effort_in_hrs_col,
get_esti_effort_in_hrs_npos,
get_time_spent_in_sec_col,
get_time_spent_in_sec_npos,
get_effort_cons_in_hrs_col,
get_effort_cons_in_hrs_npos,
get_pending_effort_in_hrs_col,
get_pending_effort_in_hrs_npos,
get_story_actual_start_col,
get_story_actual_start_npos,
get_story_actual_end_col,
get_story_actual_end_npos,
get_story_effort_comp_col,
get_story_effort_comp_npos,
get_story_schedule_progress_col,
get_story_schedule_progress_npos,
get_story_schedule_overrun_col,
get_story_schedule_overrun_npos,
get_remarks_col,
get_remarks_npos,

)

ik_pos = get_issue_key_pos()
ii_pos = get_issue_id_pos()
cfl_pos = get_custom_field_link_pos()
en_pos = get_ename_pos()
a_pos = get_assignee_pos()
oe_pos = get_orig_esti_pos() 
cfp_pos = get_custom_field_points_pos()
re_pos = get_rem_esti_pos()
ts_pos = get_time_spent_pos()
s1_pos = get_sprint_pos()
s2_pos = get_sprint2_pos()
s3_pos = get_sprint3_pos()
s_pos = get_summary_pos()
sno_pos = get_sno_pos()
epic_cfl = get_epic_custom_field_link()
epic_en_pos = get_epic_ename()
epic_es_pos = get_epic_esdate()
epic_et_pos = get_epic_etdate()
epic_eas_pos = get_epic_easDate()
epic_eae_pos = get_epic_eaedate()
ei_col = get_epic_id_col()
ei_npos = get_epic_id_npos()
si_col = get_story_id_col()
si_npos = get_story_id_npos()
sprint_id_col = get_sprint_id_col()
sprint_id_npos = get_sprint_id_npos()
sd_col = get_story_desc_col()
sd_npos = get_story_desc_npos()
at_col = get_assigned_to_col()
at_npos = get_assigned_to_npos()
esp_col = get_esti_story_points_col()
esp_npos = get_esti_story_points_npos()
sps_col = get_story_planned_start_col()
sps_npos = get_story_planned_start_npos()
spe_col = get_story_planned_end_col()
spe_npos = get_story_planned_end_npos()
ee_in_hrs_col = get_esti_effort_in_hrs_col()
ee_in_hrs_npos = get_esti_effort_in_hrs_npos()
ec_in_hrs_col = get_effort_cons_in_hrs_col()
ec_in_hrs_npos = get_effort_cons_in_hrs_npos()
ts_in_sec_col = get_time_spent_in_sec_col()
ts_in_sec_npos = get_time_spent_in_sec_npos()
pe_in_hrs_col = get_pending_effort_in_hrs_col()
pe_in_hrs_npos = get_pending_effort_in_hrs_npos()
sas_col = get_story_actual_start_col()
sas_npos = get_story_actual_start_npos()
sae_col = get_story_actual_end_col()
sae_npos = get_story_actual_end_npos()
sec_col = get_story_effort_comp_col()
sec_npos = get_story_effort_comp_npos()
ssp_col = get_story_schedule_progress_col()
ssp_npos = get_story_schedule_progress_npos()
sso_col = get_story_schedule_overrun_col()
sso_npos = get_story_schedule_overrun_npos()
r_col = get_remarks_col()
r_npos = get_remarks_npos()
def append_rows_by_name(wbook, dst_wname, src1_wname, src2_wname): 
    wb = load_workbook(wbook)
    swsheet1 = wb.get_sheet_by_name(src1_wname)
    swsheet2 = wb.get_sheet_by_name(src2_wname)

    try:
        dwsheet = wb.get_sheet_by_name(dst_wname)
        print("Worksheet '%s' found" %(dst_wname))

        print("Removing worksheet: '%s'" %(dst_wname))
        wb.remove_sheet(dwsheet)
    except:
        print("Worksheet '%s' not found" %(dst_wname))
    finally:
        print("Creating new worksheet : '%s'" %(dst_wname))
        dwsheet = wb.create_sheet(dst_wname)

    srcount1 = swsheet1.max_row
    sccount1 = swsheet1.max_column

    srcount2 = swsheet2.max_row
    sccount2 = swsheet2.max_column

    drcount = dwsheet.max_row
    dccount = dwsheet.max_column
	
    dwsheet.insert_cols(1,19)
	
    dwsheet[1][sno_pos].value = swsheet2[1][sno_pos].value
    dwsheet[1][ei_npos].value = ei_col
    dwsheet[1][sprint_id_npos].value = sprint_id_col
    dwsheet[1][si_npos].value = si_col
    dwsheet[1][sd_npos].value = sd_col
    dwsheet[1][at_npos].value = at_col
    dwsheet[1][esp_npos].value = esp_col
    dwsheet[1][sps_npos].value = sps_col
    dwsheet[1][spe_npos].value = spe_col
    dwsheet[1][ee_in_hrs_npos].value = ee_in_hrs_col
    dwsheet[1][ts_in_sec_npos].value = ts_in_sec_col
    dwsheet[1][ec_in_hrs_npos].value = ec_in_hrs_col
    dwsheet[1][pe_in_hrs_npos].value = pe_in_hrs_col
    dwsheet[1][sas_npos].value = sas_col
    dwsheet[1][sae_npos].value = sae_col
    dwsheet[1][sec_npos].value = sec_col
    dwsheet[1][ssp_npos].value = ssp_col
    dwsheet[1][sso_npos].value = sso_col
    dwsheet[1][r_npos].value = r_col
    for i in range(drcount+1, srcount2 + 1):
        dwsheet[i][sno_pos].value = swsheet2[i][sno_pos].value
        dwsheet[i][ei_npos].value = swsheet2[i][cfl_pos].value
        dwsheet[i][si_npos].value = swsheet2[i][ik_pos].value
        dwsheet[i][sd_npos].value = swsheet2[i][s_pos].value
        dwsheet[i][at_npos].value = swsheet2[i][a_pos].value
        dwsheet[i][esp_npos].value = swsheet2[i][cfp_pos].value
        dwsheet[i][ee_in_hrs_npos].value = round(int(swsheet2[i][oe_pos].value) / 3600)
        dwsheet[i][ts_in_sec_npos].value = swsheet2[i][ts_pos].value
        dwsheet[i][ec_in_hrs_npos].value = round(int(swsheet2[i][re_pos].value)/3600)
        dwsheet[i][pe_in_hrs_npos].value = round(int(dwsheet[i][ee_in_hrs_npos].value) - int(dwsheet[i][ec_in_hrs_npos].value))
        #dwsheet[i][sec_npos].value = (round(int(dwsheet[i][ec_in_hrs_npos].value)/int(dwsheet[i][ee_in_hrs_npos].value)) * 100)  
        dwsheet[i][sec_npos].value = round(int(dwsheet[i][ec_in_hrs_npos].value)/ int(dwsheet[i][ee_in_hrs_npos].value)*100)
        dwsheet[i][ssp_npos].value = "100%"
        dwsheet[i][sso_npos].value = "0%"
        
        s1 = swsheet2[i][s1_pos].value
        s2 = swsheet2[i][s2_pos].value
        s3 = swsheet2[i][s3_pos].value
        sprint = [s1, s2, s3]
        sprint = list(filter(None, sprint))
        sprint.sort(reverse = True) 
        dwsheet[i][sprint_id_npos].value = sprint[0]
    for j in range(drcount+1, srcount1+1):
        for i in range(drcount+1, srcount2+1):
            
            if (swsheet2[i][cfl_pos].value == swsheet1[j][epic_cfl].value):

                dwsheet[i][sps_npos].value = swsheet1[j][epic_es_pos].value
                dwsheet[i][spe_npos].value = swsheet1[j][epic_et_pos].value
                dwsheet[i][sas_npos].value = swsheet1[j][epic_eas_pos].value
                dwsheet[i][sae_npos].value = swsheet1[j][epic_eae_pos].value
    wb.save(wbook)


def main():
    epic_data = "01-epics"
    stories_data = "02-stories"
    wbook = "generate-files.xlsx"
    dst_wname = "stories"

    append_rows_by_name(wbook, dst_wname, epic_data, stories_data)

if (__name__ == "__main__"):
    main()
