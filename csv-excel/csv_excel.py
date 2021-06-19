import csv
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, colors, Font, Fill

from epics_stories_data import( get_issue_key_pos,
get_issue_id_pos,
get_custom_field_link_pos,
get_ename_pos,
get_assignee_pos,
get_orig_esti_pos, 
get_custom_field_points_pos,
get_rem_esti_pos,
get_time_spent_pos,
get_sprint_pos,
get_sprint2_pos,
get_sprint3_pos,
get_summary_pos,
get_sno_pos,
get_epic_custom_field_link,
get_epic_ename,
get_epic_esdate,
get_epic_etdate,
get_epic_easDate,
get_epic_eaedate,
get_epic_id_col,
get_epic_id_npos,
get_story_id_col,
get_story_id_npos,
get_sprint_id_col,
get_sprint_id_npos,
get_story_desc_col,
get_story_desc_npos,
get_assigned_to_col,
get_assigned_to_npos,
get_esti_story_points_col,
get_esti_story_points_npos,
get_story_planned_start_col,
get_story_planned_start_npos,
get_story_planned_end_col,
get_story_planned_end_npos,
get_esti_effort_in_hrs_col,
get_esti_effort_in_hrs_npos,
get_time_spent_in_sec_col,
get_time_spent_in_sec_npos,
get_effort_cons_in_hrs_col,
get_effort_cons_in_hrs_npos,
get_pending_effort_in_hrs_col,
get_pending_effort_in_hrs_npos,
get_story_actual_start_col,
get_story_actual_start_npos,
get_story_actual_end_col,
get_story_actual_end_npos,
get_story_effort_comp_col,
get_story_effort_comp_npos,
get_story_schedule_progress_col,
get_story_schedule_progress_npos,
get_story_schedule_overrun_col,
get_story_schedule_overrun_npos,
get_remarks_col,
get_remarks_npos,

)

ik_pos = get_issue_key_pos()
ii_pos = get_issue_id_pos()
cfl_pos = get_custom_field_link_pos()
en_pos = get_ename_pos()
a_pos = get_assignee_pos()
oe_pos = get_orig_esti_pos() 
cfp_pos = get_custom_field_points_pos()
re_pos = get_rem_esti_pos()
ts_pos = get_time_spent_pos()
s1_pos = get_sprint_pos()
s2_pos = get_sprint2_pos()
s3_pos = get_sprint3_pos()
s_pos = get_summary_pos()
sno_pos = get_sno_pos()
epic_cfl = get_epic_custom_field_link()
epic_en_pos = get_epic_ename()
epic_es_pos = get_epic_esdate()
epic_et_pos = get_epic_etdate()
epic_eas_pos = get_epic_easDate()
epic_eae_pos = get_epic_eaedate()
ei_col = get_epic_id_col()
ei_npos = get_epic_id_npos()
si_col = get_story_id_col()
si_npos = get_story_id_npos()
sprint_id_col = get_sprint_id_col()
sprint_id_npos = get_sprint_id_npos()
sd_col = get_story_desc_col()
sd_npos = get_story_desc_npos()
at_col = get_assigned_to_col()
at_npos = get_assigned_to_npos()
esp_col = get_esti_story_points_col()
esp_npos = get_esti_story_points_npos()
sps_col = get_story_planned_start_col()
sps_npos = get_story_planned_start_npos()
spe_col = get_story_planned_end_col()
spe_npos = get_story_planned_end_npos()
ee_in_hrs_col = get_esti_effort_in_hrs_col()
ee_in_hrs_npos = get_esti_effort_in_hrs_npos()
ec_in_hrs_col = get_effort_cons_in_hrs_col()
ec_in_hrs_npos = get_effort_cons_in_hrs_npos()
ts_in_sec_col = get_time_spent_in_sec_col()
ts_in_sec_npos = get_time_spent_in_sec_npos()
pe_in_hrs_col = get_pending_effort_in_hrs_col()
pe_in_hrs_npos = get_pending_effort_in_hrs_npos()
sas_col = get_story_actual_start_col()
sas_npos = get_story_actual_start_npos()
sae_col = get_story_actual_end_col()
sae_npos = get_story_actual_end_npos()
sec_col = get_story_effort_comp_col()
sec_npos = get_story_effort_comp_npos()
ssp_col = get_story_schedule_progress_col()
ssp_npos = get_story_schedule_progress_npos()
sso_col = get_story_schedule_overrun_col()
sso_npos = get_story_schedule_overrun_npos()
r_col = get_remarks_col()
r_npos = get_remarks_npos()

def get_file_by_name(file_name):
    data = []
    fd = open(file_name, 'rt')
    fname = csv.reader(fd)
    for row in fname:
        data.append(row)
    fd.close()
    return data

def create_workbook(wbook):
   
    try:
        wb = load_workbook(wbook)
        print("Workbook '%s'exists" %(wb))
    except:
        print("Creating worksheet: '%s'" %(wbook))
        wb = Workbook()

    wb.save(wbook)

def create_worksheet(wbook, wsheet):
    wb = load_workbook(wbook)
    try:
        dwsheet = wb.get_sheet_by_name(wsheet)
        print("Worksheet '%s' found" %(dwsheet))

        print("removing worksheet '%s'" %(dwsheet))
        wb.remove_sheet(dwsheet)
    except:
        print("Worksheet '%s' not found" %(wsheet))
    finally:
        print("Creating worksheet '%s'" %(wsheet))
        dwsheet = wb.create_sheet(wsheet)

    wb.save(wbook)
def get_col_names(sdata, wbook, wsheet):
    wb = load_workbook(wbook)
    dwsheet = wb.get_sheet_by_name(wsheet)
    
    dwsheet.insert_cols(1,19)

    dwsheet[1][sno_pos].value = sdata[0][sno_pos]
    dwsheet[1][ei_npos].value = ei_col
    dwsheet[1][sprint_id_npos].value = sprint_id_col
    dwsheet[1][si_npos].value = si_col
    dwsheet[1][sd_npos].value = sd_col
    dwsheet[1][at_npos].value = at_col
    dwsheet[1][esp_npos].value = esp_col
    dwsheet[1][sps_npos].value = sps_col
    dwsheet[1][spe_npos].value = spe_col
    dwsheet[1][ee_in_hrs_npos].value = ee_in_hrs_col
    dwsheet[1][ts_in_sec_npos].value = ts_in_sec_col
    dwsheet[1][ec_in_hrs_npos].value = ec_in_hrs_col
    dwsheet[1][pe_in_hrs_npos].value = pe_in_hrs_col
    dwsheet[1][sas_npos].value = sas_col
    dwsheet[1][sae_npos].value = sae_col
    dwsheet[1][sec_npos].value = sec_col
    dwsheet[1][ssp_npos].value = ssp_col
    dwsheet[1][sso_npos].value = sso_col
    dwsheet[1][r_npos].value = r_col

    wb.save(wbook)
def get_values_for_col(wbook, wsheet, sdata, edata):
    wb = load_workbook(wbook)
    dwsheet = wb.get_sheet_by_name(wsheet)
    e = len(edata)
    s = len(sdata)
    for i in range(1, s):
        j = i + 1
        dwsheet[j][sno_pos].value = sdata[i][sno_pos]
        dwsheet[j][ei_npos].value = sdata[i][cfl_pos]
        dwsheet[j][si_npos].value = sdata[i][ik_pos]
        dwsheet[j][sd_npos].value = sdata[i][s_pos]
        dwsheet[j][at_npos].value = sdata[i][a_pos]
        dwsheet[j][esp_npos].value = sdata[i][cfp_pos]
        dwsheet[j][ee_in_hrs_npos].value = round(int(sdata[i][oe_pos]) / 3600)
        dwsheet[j][ts_in_sec_npos].value = sdata[i][ts_pos]
        dwsheet[j][ec_in_hrs_npos].value = round(int(sdata[i][re_pos])/3600)
        dwsheet[j][pe_in_hrs_npos].value = round(int(dwsheet[j][ee_in_hrs_npos].value) - int(dwsheet[j][ec_in_hrs_npos].value))
        dwsheet[j][sec_npos].value = round(int(dwsheet[j][ec_in_hrs_npos].value)/ int(dwsheet[j][ee_in_hrs_npos].value)*100)
        dwsheet[j][ssp_npos].value = "100%"
        dwsheet[j][sso_npos].value = "0%"
        
        s1 = sdata[i][s1_pos]
        s2 = sdata[i][s2_pos]
        s3 = sdata[i][s3_pos]
        sprint = [s1, s2, s3]
        sprint = list(filter(None, sprint))
        sprint.sort(reverse = True) 
        dwsheet[j][sprint_id_npos].value = sprint[0]
        for j in range(1, e):
            for i in range(1, s):
            
                if (sdata[i][cfl_pos] == edata[j][epic_cfl]):

                    dwsheet[i+1][sps_npos].value = edata[j][epic_es_pos]
                    dwsheet[i+1][spe_npos].value = edata[j][epic_et_pos]
                    dwsheet[i+1][sas_npos].value = edata[j][epic_eas_pos]
                    dwsheet[i+1][sae_npos].value = edata[j][epic_eae_pos]
    wb.save(wbook)


def get_cell_colors_using_patternfill(wbook, wsheet):
    wb = load_workbook(wbook)
    dwsheet = wb.get_sheet_by_name(wsheet)

    rcount = dwsheet.max_row
    ccount = dwsheet.max_column
    #print(dir(PatternFill))
    
    fill_pattern = PatternFill(patternType = 'solid', fgColor = 'CCCCFF')
    for i in range(1, ccount):
        dwsheet[1][i].fill = fill_pattern
    #dwsheet['A1'].fill = fill_pattern

    wb.save(wbook)


def main():
    epics = "01-epics.csv"
    stories = "02-stories.csv"
    wbook = "generated-file.xlsx"
    create_workbook(wbook)
    wsheet = "epics_stories"
    worksheet = create_worksheet(wbook, wsheet)
    epics_data = get_file_by_name(epics)
    stories_data = get_file_by_name(stories)
    col_names = get_col_names(stories_data, wbook, wsheet)
    values = get_values_for_col(wbook, wsheet, stories_data, epics_data)
    fill_pattern = get_cell_colors_using_patternfill(wbook, wsheet)
if (__name__ == "__main__"):
    main()
